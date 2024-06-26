import os
import time
from subprocess import Popen, PIPE
from openpyxl import Workbook, load_workbook
import tempfile
import threading

# Пути к программам
teqcpath = r'C:\\Users\\SayliV\\Desktop\\TEST\\teqc.exe'
crnx2rnx_path = r'C:\\Users\\SayliV\\Desktop\\TEST\\CRX2RNX.exe'
# путь куда сохранится полученный файл
excel_file = r'C:\\Users\\SayliV\\Desktop\\TEST\\data.xlsx'
# Путь к папке с файлами
folder_path = r'C:\Users\SayliV\Desktop\2000_otrab'
# Глобальные переменные для счётчика и времени работы
processed_files_count = 0
start_time = time.time()
# Флаг для проверки, нужно ли перезаписать файл
first_run = True

# функция нахождения времени обработки 'О' файла
def find_data(teqc_output):
    # создание переменных для записи в них полученныйх данных
    start_window = None
    end_window = None
    window_length = None
    # нахождение по ключевым словам нужных данных
    for line in teqc_output.splitlines():
        if "Time of start of window" in line:
            start_window = line.split(':', 1)[1].strip()
        elif "Time of  end  of window" in line:
            end_window = line.split(':', 1)[1].strip()
        elif "Time line window length" in line:
            window_length = line.split(':', 1)[1].strip()
    return start_window, end_window, window_length

# функция записи полученных данных в эксель таблицу
def write_to_excel(file_basename, start_window, end_window, window_length, error_reason=None):
    global first_run
    if first_run or not os.path.exists(excel_file):
        wb = Workbook()
        ws = wb.active
        ws.append(
            ["Filename", "Time of Start of Window", "Time of End of Window", "Time Line Window Length", "Error Reason"])
        first_run = False
    else:
        wb = load_workbook(excel_file)
        ws = wb.active

    ws.append([file_basename, start_window, end_window, window_length, error_reason])
    wb.save(excel_file)

# функция работы в потоке с двумя программами
def process_file(filename):
    global processed_files_count
    try:
        file_basename = os.path.basename(filename)

        # Создаем процесс для crnx2rnx
        crnx2 = Popen([crnx2rnx_path, '-', filename], stdout=PIPE, stderr=PIPE)
        crnx2_stdout, crnx2_stderr = crnx2.communicate()

        if crnx2.returncode != 0:
            error_reason = f"CRX2RNX Error: {crnx2_stderr.decode().strip()}"
            print(f"Error in CRX2RNX for {file_basename}: {error_reason}")
            write_to_excel(file_basename, None, None, None, error_reason)
            return

        # Используем временный файл для промежуточного хранения данных
        with tempfile.NamedTemporaryFile(delete=False) as temp_file:
            intermediate_file = temp_file.name
            temp_file.write(crnx2_stdout)

        # Создаем процесс для teqc, используя промежуточный файл
        teqc = Popen([teqcpath, '+qc', intermediate_file], stdout=PIPE, stderr=PIPE)
        teqc_stdout, teqc_stderr = teqc.communicate()

        if teqc.returncode != 0:
            error_reason = f"TEQC Error: {teqc_stderr.decode().strip()}"
            print(f"Error in TEQC for {file_basename}: {error_reason}")
            write_to_excel(file_basename, None, None, None, error_reason)
            return

        teqc_output_decoded = teqc_stdout.decode()
        start_window, end_window, window_length = find_data(teqc_output_decoded)

        if start_window and end_window and window_length:
            write_to_excel(file_basename, start_window, end_window, window_length)
            processed_files_count += 1
        else:
            error_reason = "TEQC Error: Failed to parse required data from TEQC output."
            print(f"{error_reason} for {file_basename}.")
            write_to_excel(file_basename, None, None, None, error_reason)

    except Exception as e:
        error_reason = f"General Error: {e}"
        print(f"An error occurred while processing {filename}: {error_reason}")
        write_to_excel(file_basename, None, None, None, error_reason)

    finally:
        # Удаляем промежуточный файл
        if 'intermediate_file' in locals() and os.path.exists(intermediate_file):
            os.remove(intermediate_file)

def print_status():
    while True:
        elapsed_time = time.time() - start_time
        print(f"Processed files: {processed_files_count}, Elapsed time: {elapsed_time:.2f} seconds")
        time.sleep(10)

# Проверяем и удаляем существующий Excel файл перед началом работы
if os.path.exists(excel_file):
    os.remove(excel_file)

# Запускаем поток для вывода статуса каждые 10 секунд
status_thread = threading.Thread(target=print_status)
status_thread.daemon = True  # Поток завершится при завершении основного потока
status_thread.start()

# Обрабатываем все файлы в указанной папке вне зависимости от их расширения
for root, dirs, files in os.walk(folder_path):
    for file in files:
        process_file(os.path.join(root, file))

# Если необзходимо обработать файлы определенного года
# for root, dirs, files in os.walk(folder_path):
#     for file in files:
#         if file.endswith('.00d'):
#             process_file(os.path.join(root, file))

# Ожидаем завершения всех потоков перед выходом из программы
status_thread.join(0)
